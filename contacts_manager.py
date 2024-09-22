import openpyxl
from openpyxl import Workbook, load_workbook
import os

# Define a class to represent a Contact
class Contact:
    def __init__(self, name, phone, email):
        self.name = name
        self.phone = phone
        self.email = email

    def __str__(self):
        return f"Name: {self.name}, Phone: {self.phone}, Email: {self.email}"

# Define a class to manage the list of contacts
class ContactList:
    def __init__(self):
        self.contacts = []
        self.file_name = 'contacts_db.xlsx'
        self.load_contacts()  # Load contacts from Excel file when initializing

    # Add a new contact to the list and save it to the Excel file
    def add_contact(self, contact):
        self.contacts.append(contact)
        self.save_contacts()

    # Delete a contact by name and save the changes to the Excel file
    def delete_contact(self, name):
        self.contacts = [contact for contact in self.contacts if contact.name != name]
        self.save_contacts()

    # Search for a contact by name
    def search_contact(self, name):
        for contact in self.contacts:
            if contact.name == name:
                return contact
        return None

    # Display all contacts
    def display_contacts(self):
        for contact in self.contacts:
            print(contact)

    # Open the Excel file containing the contacts
    def open_excel_file(self):
        try:
            os.startfile(self.file_name)  # For Windows
        except AttributeError:
            os.system(f'open {self.file_name}')  # For macOS
        except Exception as e:
            print(f"Unable to open Excel file: {e}")

    # Load contacts from the Excel file
    def load_contacts(self):
        try:
            workbook = load_workbook(self.file_name)
            sheet = workbook.active
            # Iterate over the rows in the Excel sheet and create Contact objects
            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, phone, email = row
                self.contacts.append(Contact(name, phone, email))
            workbook.close()
        except FileNotFoundError:
            # If the file doesn't exist, create a new one with headers
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Name", "Phone", "Email"])
            workbook.save(self.file_name)
            workbook.close()

    # Save the contacts to the Excel file
    def save_contacts(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Phone", "Email"])
        for contact in self.contacts:
            sheet.append([contact.name, contact.phone, contact.email])
        workbook.save(self.file_name)
        workbook.close()

# Validate the phone number to ensure it is 10 digits long
def validate_phone(phone):
    if len(phone) == 10 and phone.isdigit():
        return phone
    else:
        print("Invalid phone number. Please enter a 10-digit phone number.")
        new_phone = input("Enter phone: ")
        return validate_phone(new_phone)

# Main function to interact with the user
def main():
    contact_list = ContactList()

    while True:
        print("\nContactMaster")
        print("1. Add Contact")
        print("2. Delete Contact")
        print("3. Search Contact")
        print("4. Display All Contacts")
        print("5. Exit")
        choice = input("Enter your choice: ")

        if choice == '1':
            name = input("Enter name: ")
            phone = input("Enter phone: ")
            phone = validate_phone(phone)
            email = input("Enter email: ")
            contact = Contact(name, phone, email)
            contact_list.add_contact(contact)

        elif choice == '2':
            name = input("Enter the name of the contact to delete: ")
            contact = contact_list.search_contact(name)
            if contact:
                contact_list.delete_contact(name)
                print(f'Contact {name} deleted.')
            else:
                print(f'Contact {name} not found.')

        elif choice == '3':
            name = input("Enter the name of the contact to search: ")
            contact = contact_list.search_contact(name)
            if contact:
                print(contact)
            else:
                print(f'Contact {name} not found.')

        elif choice == '4':
            contact_list.display_contacts()
            contact_list.open_excel_file()

        elif choice == '5':
            contact_list.save_contacts()
            break

        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
